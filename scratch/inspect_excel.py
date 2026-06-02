import openpyxl

path = "/Users/ismailcan/Downloads/Çanakkale_Ezine_Çamlıca ruhsatı_ Numune Teslim Formu OCAK İCİ.xlsx"

try:
    wb = openpyxl.load_workbook(path, read_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print("Max rows in sheet:", sheet.max_row)
        print("Max columns in sheet:", sheet.max_column)
        # Print every non-empty row in the sheet
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if any(val is not None for val in row):
                non_empty = {col_idx: val for col_idx, val in enumerate(row) if val is not None}
                print(f"Row {idx+1}: {non_empty}")
except Exception as e:
    print("Error reading Excel:", e)
