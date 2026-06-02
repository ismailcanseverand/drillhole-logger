import openpyxl
import json
import os
import re
import datetime

# Helper to normalize kuyu/sondaj names
def clean_id(val):
    if val is None:
        return ""
    return str(val).strip().upper()

# Robust string-to-float converter
def to_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except ValueError:
        m = re.findall(r"[-+]?\d*\.?\d+", str(val))
        if m:
            try:
                return float(m[0])
            except:
                pass
        return default

# Known loggers for checking logger vs project names
KNOWN_LOGGERS = ["İsmailcan SEVER", "Altan COŞKUN", "Levent CAN", "Mehmet KOLDANCI", "Muhammed KAYALIDAĞ", "Mustafa KAŞ", "Emir Özçakıcı"]

def map_lithology_to_code(description):
    desc = str(description).lower().strip()
    # Normalize Turkish characters
    desc = desc.replace('ı', 'i').replace('ö', 'o').replace('ü', 'u').replace('ş', 's').replace('ç', 'c').replace('ğ', 'g')
    desc = desc.replace('i̇', 'i')
    
    if not desc:
        return 'GNAYS'
    if 'dolgu' in desc or 'pasa' in desc:
        return 'DOLGU'
    if 'albit' in desc or 'feldspat' in desc or 'feldispat' in desc:
        return 'ALBIT'
    if 'gnays' in desc:
        return 'GNAYS'
    if 'andezit' in desc or 'tuf' in desc or 'tüf' in desc:
        return 'ANDEZIT'
    if 'kaolen' in desc or 'kaolin' in desc:
        return 'KAOLEN'
    if 'granit' in desc or 'grani' in desc:
        return 'GRANIT'
    if 'kuvars' in desc or 'sileks' in desc or 'kuvarsit' in desc:
        return 'KUVARSIT'
    if 'sist' in desc or 'şist' in desc:
        return 'SIST'
    if 'kil' in desc:
        return 'KIL'
    if 'kum' in desc:
        return 'KUM'
    if 'komur' in desc or 'kömür' in desc:
        return 'KOMUR'
    if 'bres' in desc or 'breş' in desc or 'fay' in desc:
        return 'BRES'
    if 'halloysit' in desc:
        return 'HALLOYSIT'
    if 'kalsit' in desc or 'kirec' in desc or 'kalk' in desc:
        return 'KALSIT'
    if 'oksit' in desc or 'oksi' in desc:
        return 'OKSIT'
    if 'sulfit' in desc or 'sülfi' in desc:
        return 'SULFIT'
    if 'ignimb' in desc:
        return 'IGNIMBIRIT'
    if 'intruzif' in desc or 'intruzi' in desc:
        return 'INTRUZIF'
    if 'perlit' in desc:
        return 'PERLIT'
    if 'dasit' in desc:
        return 'DASIT'
    if 'riyolit' in desc or 'riyo' in desc:
        return 'RIYOLIT'
    if 'dayk' in desc:
        return 'DAYK'
    if 'volkanosed' in desc:
        return 'VOLKANOSEDIMANTER'
    if 'siyenit' in desc or 'syenit' in desc:
        return 'SIYENIT'
    if 'granod' in desc:
        return 'GRANODIYORIT'
    if 'alunit' in desc:
        return 'ALUNIT'
    if 'toprak' in desc:
        return 'TOPRAK'
    return 'GNAYS'

# Load existing JSON database
json_path = 'public/drillhole_data.json'
compiled_data = {}
if os.path.exists(json_path):
    print("Loading existing JSON database...")
    with open(json_path, 'r', encoding='utf-8') as f:
        compiled_data = json.load(f)
    print(f"Loaded {len(compiled_data)} existing industrial drillholes.")

# 1. Clean and migrate existing industrial drillholes project name
for h_id, h_data in compiled_data.items():
    col = h_data.get('collar', {})
    logger_val = col.get('logger', '')
    
    # If logger is actually a project name (e.g. CEYHAN, MILAS)
    if logger_val and logger_val not in KNOWN_LOGGERS:
        col['project'] = logger_val
        col['logger'] = ""
    else:
        if 'project' not in col:
            col['project'] = ""
    
    h_data['collar'] = col

# 2. Open new metallic Excel file
excel_path = '/Users/ismailcan/Downloads/KM-Metalik_Sondaj_All.xlsx'
print(f"Parsing new metallic Excel file at {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

m_collars = {}
m_surveys = {}
m_lithologies = {}
m_geotechs = {}
m_assays = {}

# A. Parse Collars
sheet_collar = wb['DH_Collar']
for row in sheet_collar.iter_rows(min_row=2, values_only=True):
    hole_id, easting, northing, elevation, azimuth, dip, depth, project, license, start = row[:10]
    if not hole_id:
        continue
    c_id = clean_id(hole_id)
    
    start_date = ""
    if isinstance(start, datetime.datetime):
        start_date = start.strftime("%Y-%m-%d")
    elif start:
        start_date = str(start)[:10]

    m_collars[c_id] = {
        "holeId": str(hole_id).strip(),
        "easting": to_float(easting, 0.0),
        "northing": to_float(northing, 0.0),
        "elevation": to_float(elevation, 0.0),
        "totalDepth": to_float(depth, 0.0),
        "dip": to_float(dip, -90.0),
        "azimuth": to_float(azimuth, 0.0),
        "dateStarted": start_date,
        "dateCompleted": "",
        "logger": "",
        "status": "Completed",
        "project": str(project).strip() if project else ""
    }

# B. Parse Surveys
sheet_survey = wb['DownHole_Survey']
for r_idx, row in enumerate(sheet_survey.iter_rows(min_row=2, values_only=True), start=2):
    hole_id, depth, meas_az, dip, mag_dev, real_az, desc = row[:7]
    if not hole_id:
        continue
    c_id = clean_id(hole_id)
    if c_id not in m_surveys:
        m_surveys[c_id] = []
        
    azimuth_val = real_az if real_az is not None else (meas_az if meas_az is not None else 0.0)
    m_surveys[c_id].append({
        "id": f"s-{r_idx}",
        "depth": to_float(depth, 0.0),
        "dip": to_float(dip, -90.0),
        "azimuth": to_float(azimuth_val, 0.0)
    })

# C. Parse Lithology
sheet_litho = wb['DH_Lithology']
for r_idx, row in enumerate(sheet_litho.iter_rows(min_row=2, values_only=True), start=2):
    proje, hole_id, bas, bit, aralik, lito_code, litoloji = row[:7]
    if not hole_id:
        continue
    c_id = clean_id(hole_id)
    if c_id not in m_lithologies:
        m_lithologies[c_id] = []
        
    m_lithologies[c_id].append({
        "id": f"l-{r_idx}",
        "from": to_float(bas, 0.0),
        "to": to_float(bit, 0.0),
        "rockCode": str(lito_code).strip().upper() if lito_code else map_lithology_to_code(litoloji),
        "alteration": "",
        "mineralization": "",
        "description": str(litoloji).strip() if litoloji else ""
    })

# D. Parse Geotech Core Runs (TCR)
sheet_tcr = wb['TCR']
for r_idx, row in enumerate(sheet_tcr.iter_rows(min_row=2, values_only=True), start=2):
    proje, hole_id, m_den, m_ye, aralik, man_karot, cr, cm_karot, rqd = row[:9]
    if not hole_id:
        continue
    c_id = clean_id(hole_id)
    if c_id not in m_geotechs:
        m_geotechs[c_id] = []
        
    m_geotechs[c_id].append({
        "id": f"g-{r_idx}",
        "from": to_float(m_den, 0.0),
        "to": to_float(m_ye, 0.0),
        "drilledLength": to_float(aralik, to_float(m_ye, 0.0) - to_float(m_den, 0.0)),
        "recoveredLength": to_float(man_karot, 0.0),
        "solidPiecesOver10cm": to_float(cm_karot, 0.0),
        "tcrPercent": to_float(cr, 0.0),
        "rqdPercent": to_float(rqd, 0.0)
    })

# E. Parse DH_Analysis
sheet_anal = wb['DH_Analysis']
for r_idx, row in enumerate(sheet_anal.iter_rows(min_row=2, values_only=True), start=2):
    hole_id, proj, lic, sample_id, bas, bit, interval, lito = row[:8]
    if not hole_id or not sample_id:
        continue
    c_id = clean_id(hole_id)
    if c_id not in m_assays:
        m_assays[c_id] = []
        
    m_assays[c_id].append({
        "id": f"a-{r_idx}",
        "sampleId": str(sample_id).strip(),
        "from": to_float(bas, 0.0),
        "to": to_float(bit, 0.0),
        "sampleType": "Core",
        "al2o3": 0.0,
        "fe2o3": 0.0,
        "sio2": 0.0,
        "tio2": 0.0,
        "na2o_k2o": 0.0,
        "loi": 0.0
    })

# 3. Merge metallic data arrays into compiled_data
merged_count = 0
for c_id, collar in m_collars.items():
    # Gather surveys
    surveys = m_surveys.get(c_id, [])
    if not surveys:
        # Default top and bottom surveys
        surveys = [
            {"id": "s1", "depth": 0.0, "dip": collar["dip"], "azimuth": collar["azimuth"]},
            {"id": "s2", "depth": collar["totalDepth"], "dip": collar["dip"], "azimuth": collar["azimuth"]}
        ]
    else:
        # Sort by depth
        surveys.sort(key=lambda s: s["depth"])
        
    litho = m_lithologies.get(c_id, [])
    litho.sort(key=lambda l: l["from"])
    
    geotech = m_geotechs.get(c_id, [])
    geotech.sort(key=lambda g: g["from"])
    
    assays = m_assays.get(c_id, [])
    assays.sort(key=lambda a: a["from"])
    
    compiled_data[c_id] = {
        "collar": collar,
        "surveys": surveys,
        "lithology": litho,
        "geotech": geotech,
        "assays": assays
    }
    merged_count += 1

print(f"Successfully compiled. Writing results back to {json_path}...")
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(compiled_data, f, ensure_ascii=False, indent=2)

print(f"Merge Complete! Total active drillholes in JSON database: {len(compiled_data)} (Merged {merged_count} new metallic holes).")
