import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill

file_path = 'public/Numune_Teslim_Formu.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 1. Set SO4 (Col 25, Row 9) bottom border to medium for header
cell_so4 = ws.cell(row=9, column=25)
existing_border = cell_so4.border
header_border = Border(
    left=existing_border.left,
    right=existing_border.right,
    top=existing_border.top,
    bottom=Side(style='medium')
)
cell_so4.border = header_border

# 2. Add border to all cells in SO4 column (rows 10-47)
for r in range(10, 48):
    cell = ws.cell(row=r, column=25)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# 3. Set Düşünceler (Col 30, Row 9) alignment textRotation to 90 (vertical)
cell_desc = ws.cell(row=9, column=30)
cell_desc.alignment = Alignment(horizontal='center', vertical='center', textRotation=90)

# 4. Apply thin borders around all cells in Düşünceler column (rows 10-47)
for r in range(10, 48):
    cell = ws.cell(row=r, column=30)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# 5. Save workbook
wb.save(file_path)
print("Template styles updated successfully.")
