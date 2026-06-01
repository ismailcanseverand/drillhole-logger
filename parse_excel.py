import openpyxl
import json
import os
import re

wb = openpyxl.load_workbook('/Users/ismailcan/Ideas/KaleMaden/Endüstriyel_Sondaj-24-06-07.xlsx', data_only=True)

# Helper to normalize kuyu/sondaj names
def clean_id(val):
    if val is None:
        return ""
    return str(val).strip().upper()

# Robust string-to-float converter
def to_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except ValueError:
        # Extract first numeric value found in the string
        m = re.findall(r"[-+]?\d*\.?\d+", str(val))
        if m:
            try:
                return float(m[0])
            except:
                pass
        return default

collars = {}
surveys = {}
lithology = {}
geotech = {}
assays = {}

# 1. Parse Sondaj (Collar)
sondaj_sheet = wb['Sondaj']
for r_idx, row in enumerate(sondaj_sheet.iter_rows(min_row=2, values_only=True), start=2):
    saha, sondaj_no, dogu, kuzey, kot, az, egim, der, yil, topo = row[:10]
    if not sondaj_no:
        continue
    c_id = clean_id(sondaj_no)
    collars[c_id] = {
        "holeId": str(sondaj_no).strip(),
        "easting": to_float(dogu, 0.0),
        "northing": to_float(kuzey, 0.0),
        "elevation": to_float(kot, 0.0),
        "totalDepth": to_float(der, 0.0),
        "dip": to_float(egim, -90.0),
        "azimuth": to_float(az, 0.0),
        "dateStarted": "",
        "dateCompleted": str(yil) if yil is not None else "",
        "logger": str(saha).strip() if saha else "KaleMaden",
        "status": "Completed"
    }

def map_lithology_to_code(description):
    desc = str(description).lower().strip()
    # Normalize Turkish characters
    desc = desc.replace('ı', 'i').replace('ö', 'o').replace('ü', 'u').replace('ş', 's').replace('ç', 'c').replace('ğ', 'g')
    desc = desc.replace('i̇', 'i')
    
    if not desc:
        return 'GNAYS'
    if 'dolgu' in desc or 'pasa' in desc:
        return 'DOLGU'
    if 'albit' in desc or 'feldspat' in desc or 'feldispat' in desc:
        return 'ALBIT'
    if 'gnays' in desc:
        return 'GNAYS'
    if 'andezit' in desc or 'tuf' in desc or 'tüf' in desc:
        return 'ANDEZIT'
    if 'kaolen' in desc or 'kaolin' in desc:
        return 'KAOLEN'
    if 'granit' in desc or 'grani' in desc:
        return 'GRANIT'
    if 'kuvars' in desc or 'sileks' in desc or 'kuvarsit' in desc:
        return 'KUVARSIT'
    if 'sist' in desc or 'şist' in desc:
        return 'SIST'
    if 'kil' in desc:
        return 'KIL'
    if 'kum' in desc:
        return 'KUM'
    if 'komur' in desc or 'kömür' in desc:
        return 'KOMUR'
    if 'bres' in desc or 'breş' in desc or 'fay' in desc:
        return 'BRES'
    if 'halloysit' in desc:
        return 'HALLOYSIT'
    if 'kalsit' in desc or 'kirec' in desc or 'kalk' in desc:
        return 'KALSIT'
    if 'oksit' in desc or 'oksi' in desc:
        return 'OKSIT'
    if 'sulfit' in desc or 'sülfi' in desc:
        return 'SULFIT'
    if 'ignimb' in desc:
        return 'IGNIMBIRIT'
    if 'intruzif' in desc or 'intruzi' in desc:
        return 'INTRUZIF'
    if 'perlit' in desc:
        return 'PERLIT'
    if 'dasit' in desc:
        return 'DASIT'
    if 'riyolit' in desc or 'riyo' in desc:
        return 'RIYOLIT'
    if 'dayk' in desc:
        return 'DAYK'
    if 'volkanosed' in desc:
        return 'VOLKANOSEDIMANTER'
    if 'siyenit' in desc or 'syenit' in desc:
        return 'SIYENIT'
    if 'granod' in desc:
        return 'GRANODIYORIT'
    if 'alunit' in desc:
        return 'ALUNIT'
    if 'toprak' in desc:
        return 'TOPRAK'
    
    words = re.findall(r'[a-z]+', desc)
    if words:
        w = words[0].upper()
        if w in ['GRI', 'KIRMIZIMSI', 'YESILIMSI', 'BEJ', 'KREM', 'ACIK', 'KAHVERENGIMSI', 'BEYAZ', 'SARIMSI', 'KOYU', 'BEYAZIMSI', 'SARI', 'YESIL', 'SIYAH', 'INCE']:
            if len(words) > 1:
                return words[1].upper()
        return w
    return 'GNAYS'

# 2. Parse Litoloji
lito_sheet = wb['Litoloji']
for r_idx, row in enumerate(lito_sheet.iter_rows(min_row=2, values_only=True), start=2):
    proje, sondaj_no, bas, bit, lito_code, n1, n2, litoloji, n3 = row[:9]
    if not sondaj_no:
        continue
    c_id = clean_id(sondaj_no)
    if c_id not in lithology:
        lithology[c_id] = []
    
    from_val = to_float(bas, 0.0)
    to_val = to_float(bit, 0.0)
        
    lithology[c_id].append({
        "id": f"l-{r_idx}",
        "from": from_val,
        "to": to_val,
        "rockCode": map_lithology_to_code(litoloji),
        "alteration": "",
        "mineralization": "",
        "description": str(litoloji).strip() if litoloji else ""
    })

# 3. Parse TCR
tcr_sheet = wb['TCR']
for r_idx, row in enumerate(tcr_sheet.iter_rows(min_row=2, values_only=True), start=2):
    saha, kuyu_no, m_den, m_ye, aralik, man_karot, cr, cm_karot, rqd = row[:9]
    if not kuyu_no:
        continue
    c_id = clean_id(kuyu_no)
    if c_id not in geotech:
        geotech[c_id] = []
    
    from_val = to_float(m_den, 0.0)
    to_val = to_float(m_ye, 0.0)
    drilled = to_float(aralik, to_val - from_val)
    rec = to_float(man_karot, 0.0)
    solids = to_float(cm_karot, 0.0)
    
    tcr_val = to_float(cr, 0.0)
    rqd_val = to_float(rqd, 0.0)
        
    geotech[c_id].append({
        "id": f"g-{r_idx}",
        "from": from_val,
        "to": to_val,
        "drilledLength": drilled,
        "recoveredLength": rec,
        "solidPiecesOver10cm": solids,
        "tcrPercent": round(tcr_val, 2),
        "rqdPercent": round(rqd_val, 2)
    })

# Helper to process Assay rows
def add_assay_row(sondaj_no, numune_no, bas, bit, az, sio2, al2o3, tio2, fe2o3, cao, mgo, na2o, k2o, na2o_k2o, r_idx, source_id):
    c_id = clean_id(sondaj_no)
    if c_id not in assays:
        assays[c_id] = []
        
    from_val = to_float(bas, 0.0)
    to_val = to_float(bit, 0.0)
        
    assays[c_id].append({
        "id": f"a-{source_id}-{r_idx}",
        "sampleId": str(numune_no).strip() if numune_no else f"S-{source_id}-{r_idx}",
        "from": from_val,
        "to": to_val,
        "sampleType": "Core",
        "al2o3": to_float(al2o3, 0.0),
        "fe2o3": to_float(fe2o3, 0.0),
        "sio2": to_float(sio2, 0.0),
        "tio2": to_float(tio2, 0.0),
        "na2o_k2o": to_float(na2o_k2o, to_float(na2o, 0.0) + to_float(k2o, 0.0)),
        "loi": to_float(az, 0.0)
    })

# 4. Parse Kil_Analiz
kil_sheet = wb['Kil_Analiz']
for r_idx, row in enumerate(kil_sheet.iter_rows(min_row=2, values_only=True), start=2):
    saha, sondaj_no, num_no, bas, bit, az, sio2, al2o3, tio2, fe2o3, cao, mgo, na2o, k2o, so4, na2o_k2o, ti_fe = row[:17]
    if not sondaj_no:
        continue
    add_assay_row(sondaj_no, num_no, bas, bit, az, sio2, al2o3, tio2, fe2o3, cao, mgo, na2o, k2o, na2o_k2o, r_idx, "KIL")

# 5. Parse Aydın_Analiz
ayd_sheet = wb['Aydın_Analiz']
for r_idx, row in enumerate(ayd_sheet.iter_rows(min_row=2, values_only=True), start=2):
    saha, sondaj_no, num_no, bas, bit, az, sio2, al2o3, tio2, fe2o3, cao, mgo, na2o, k2o, na2o_k2o = row[:15]
    if not sondaj_no:
        continue
    add_assay_row(sondaj_no, num_no, bas, bit, az, sio2, al2o3, tio2, fe2o3, cao, mgo, na2o, k2o, na2o_k2o, r_idx, "AYD")

# Merge and create final JSON
all_holes = sorted(list(set(collars.keys())))
compiled_data = {}

for h_id in all_holes:
    # Ensure there is at least some logging data to include the hole
    h_litho = lithology.get(h_id, [])
    h_geotech = geotech.get(h_id, [])
    h_assays = assays.get(h_id, [])
    
    # We always generate surveys dynamically if they are empty
    h_collar = collars[h_id]
    h_surveys = [
        {"id": "s1", "depth": 0.0, "dip": h_collar["dip"], "azimuth": h_collar["azimuth"]},
        {"id": "s2", "depth": h_collar["totalDepth"], "dip": h_collar["dip"], "azimuth": h_collar["azimuth"]}
    ]
    
    compiled_data[h_id] = {
        "collar": h_collar,
        "surveys": h_surveys,
        "lithology": h_litho,
        "geotech": h_geotech,
        "assays": h_assays
    }

# Save compilation
output_dir = '/Users/ismailcan/.gemini/antigravity/scratch/drillhole-logger/public'
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, 'drillhole_data.json')

with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(compiled_data, f, ensure_ascii=False, indent=2)

print(f"Compilation finished. Total holes: {len(compiled_data)}")

# Find holes that have data in all sheets (Collar, Litoloji, TCR, Assays) to recommend as loaded demo holes
well_populated = []
for h_id, data in compiled_data.items():
    if len(data["lithology"]) > 2 and len(data["geotech"]) > 2 and len(data["assays"]) > 2:
        well_populated.append((h_id, len(data["lithology"]), len(data["geotech"]), len(data["assays"])))

print("Top 10 well populated holes:")
for x in sorted(well_populated, key=lambda item: -item[1])[:10]:
    print(f"Hole ID: {x[0]} - Litho rows: {x[1]} - TCR rows: {x[2]} - Assay rows: {x[3]}")
