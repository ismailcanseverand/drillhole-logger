import openpyxl

file_path = 'public/Numune_Teslim_Formu.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 1. Update headers for columns E (5) and F (6)
ws.cell(row=9, column=5).value = 'BAŞLANGIÇ'
ws.cell(row=9, column=6).value = 'BİTİŞ'

# 2. Delete Column G (7)
ws.delete_cols(7)

# 3. Save the modified workbook
wb.save(file_path)
print("Template updated and saved successfully.")
